import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from database import get_db, Usuario, Sessao, RecomendacaoIA
from routers.auth import get_current_user, require_profissional

router = APIRouter()


@router.get("/pdf/{atleta_id}")
def gerar_pdf(
    atleta_id: int,
    prof: Usuario = Depends(require_profissional),
    db: Session = Depends(get_db),
):
    atleta = db.query(Usuario).filter(
        Usuario.id == atleta_id, Usuario.profissional_id == prof.id
    ).first()
    if not atleta:
        raise HTTPException(status_code=404, detail="Atleta não encontrado")

    sessoes = (
        db.query(Sessao)
        .filter(Sessao.atleta_id == atleta_id, Sessao.status == "concluida")
        .order_by(desc(Sessao.criado_em))
        .all()
    )

    pdf_bytes = _build_pdf(atleta, sessoes, prof)
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_{atleta.nome.replace(' ','_')}.pdf"},
    )


@router.get("/excel/{atleta_id}")
def exportar_excel(
    atleta_id: int,
    prof: Usuario = Depends(require_profissional),
    db: Session = Depends(get_db),
):
    atleta = db.query(Usuario).filter(
        Usuario.id == atleta_id, Usuario.profissional_id == prof.id
    ).first()
    if not atleta:
        raise HTTPException(status_code=404, detail="Atleta não encontrado")

    sessoes = (
        db.query(Sessao)
        .filter(Sessao.atleta_id == atleta_id, Sessao.status == "concluida")
        .order_by(desc(Sessao.criado_em))
        .all()
    )

    xlsx_bytes = _build_excel(atleta, sessoes)
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=dados_{atleta.nome.replace(' ','_')}.xlsx"},
    )


@router.get("/dashboard-stats")
def dashboard_stats(
    prof: Usuario = Depends(require_profissional),
    db: Session = Depends(get_db),
):
    atletas = db.query(Usuario).filter(
        Usuario.profissional_id == prof.id, Usuario.ativo == True
    ).all()
    ids = [a.id for a in atletas]

    if not ids:
        return {"total_atletas": 0, "taxa_media": None, "perda_media": None, "total_sessoes": 0, "alertas": []}

    sessoes = db.query(Sessao).filter(
        Sessao.atleta_id.in_(ids), Sessao.status == "concluida"
    ).all()

    taxas = [s.taxa_sudorese for s in sessoes if s.taxa_sudorese]
    perdas = [s.variacao_peso_pct for s in sessoes if s.variacao_peso_pct]

    alertas = []
    for s in sessoes:
        if s.variacao_peso_pct and s.variacao_peso_pct > 3:
            a = db.query(Usuario).filter(Usuario.id == s.atleta_id).first()
            alertas.append({
                "tipo": "desidratacao_grave",
                "atleta_nome": a.nome if a else "Desconhecido",
                "variacao_pct": s.variacao_peso_pct,
                "sessao_id": s.id,
                "data": s.criado_em.isoformat() if s.criado_em else None,
            })

    modalidade_map: dict[str, list] = {}
    for a in atletas:
        mod = a.modalidade or "Outro"
        taxas_atleta = [s.taxa_sudorese for s in sessoes if s.atleta_id == a.id and s.taxa_sudorese]
        if taxas_atleta:
            modalidade_map.setdefault(mod, []).extend(taxas_atleta)

    por_modalidade = [
        {"modalidade": k, "taxa_media": round(sum(v) / len(v), 2)}
        for k, v in modalidade_map.items()
    ]

    return {
        "total_atletas": len(atletas),
        "taxa_media_l_h": round(sum(taxas) / len(taxas), 2) if taxas else None,
        "perda_media_pct": round(sum(perdas) / len(perdas), 2) if perdas else None,
        "total_sessoes": len(sessoes),
        "por_modalidade": por_modalidade,
        "alertas": sorted(alertas, key=lambda x: x["data"] or "", reverse=True)[:10],
    }



def _build_pdf(atleta: Usuario, sessoes: list, prof: Usuario) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    crimson = colors.HexColor("#9B1C2E")

    title_style = ParagraphStyle("title", parent=styles["Heading1"], textColor=crimson, fontSize=18)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], textColor=colors.gray, fontSize=10)

    story = [
        Paragraph("Nutri-Esportiva", title_style),
        Paragraph(f"Relatório de Hidratação — {atleta.nome}", styles["Heading2"]),
        Paragraph(f"Modalidade: {atleta.modalidade or 'N/D'} | Gerado por: {prof.nome}", sub_style),
        Spacer(1, 0.5*cm),
    ]

    headers = ["Data", "Duração (min)", "Peso Pré (kg)", "Peso Pós (kg)", "Ingestão (ml)", "Taxa (L/h)", "Variação (%)"]
    rows = [headers]
    for s in sessoes:
        rows.append([
            s.criado_em.strftime("%d/%m/%Y") if s.criado_em else "—",
            f"{s.duracao_minutos:.0f}" if s.duracao_minutos else "—",
            f"{s.peso_pre:.1f}" if s.peso_pre else "—",
            f"{s.peso_pos:.1f}" if s.peso_pos else "—",
            f"{s.ingestao_ml:.0f}" if s.ingestao_ml else "—",
            f"{s.taxa_sudorese:.2f}" if s.taxa_sudorese else "—",
            f"{s.variacao_peso_pct:.1f}%" if s.variacao_peso_pct else "—",
        ])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), crimson),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F5F5")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E0D8D8")),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(table)

    doc.build(story)
    return buf.getvalue()


def _build_excel(atleta: Usuario, sessoes: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_hist = wb.active
    ws_hist.title = "Histórico"

    crimson = "9B1C2E"
    fill_header = PatternFill("solid", fgColor=crimson)
    font_header = Font(color="FFFFFF", bold=True)
    thin = Border(
        left=Side(style="thin", color="E0D8D8"),
        right=Side(style="thin", color="E0D8D8"),
        top=Side(style="thin", color="E0D8D8"),
        bottom=Side(style="thin", color="E0D8D8"),
    )

    headers = ["Data", "Duração (min)", "Peso Pré (kg)", "Peso Pós (kg)",
               "Ingestão (ml)", "Taxa Sudorese (L/h)", "Variação Peso (%)"]

    for col_idx, h in enumerate(headers, 1):
        cell = ws_hist.cell(row=1, column=col_idx, value=h)
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
        ws_hist.column_dimensions[get_column_letter(col_idx)].width = 18

    for row_idx, s in enumerate(sessoes, 2):
        row_data = [
            s.criado_em.strftime("%d/%m/%Y") if s.criado_em else None,
            s.duracao_minutos,
            s.peso_pre,
            s.peso_pos,
            s.ingestao_ml,
            s.taxa_sudorese,
            s.variacao_peso_pct,
        ]
        fill = PatternFill("solid", fgColor="FFF5F5") if row_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for col_idx, val in enumerate(row_data, 1):
            cell = ws_hist.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            cell.fill = fill
            cell.border = thin

    ws_stats = wb.create_sheet("Estatísticas")
    taxas = [s.taxa_sudorese for s in sessoes if s.taxa_sudorese]
    perdas = [s.variacao_peso_pct for s in sessoes if s.variacao_peso_pct]

    stats = [
        ("Atleta", atleta.nome),
        ("Modalidade", atleta.modalidade or "N/D"),
        ("Total de Sessões", len(sessoes)),
        ("Taxa Média (L/h)", round(sum(taxas)/len(taxas), 2) if taxas else "N/D"),
        ("Taxa Máxima (L/h)", round(max(taxas), 2) if taxas else "N/D"),
        ("Maior Perda (%)", round(max(perdas), 2) if perdas else "N/D"),
        ("Desvio Padrão Taxa", round(_stdev(taxas), 3) if len(taxas) > 1 else "N/D"),
    ]

    for r, (label, value) in enumerate(stats, 1):
        ws_stats.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_stats.cell(row=r, column=2, value=value)
    ws_stats.column_dimensions["A"].width = 25
    ws_stats.column_dimensions["B"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stdev(values: list) -> float:
    if len(values) < 2:
        return 0.0
    n = len(values)
    mean = sum(values) / n
    return (sum((x - mean) ** 2 for x in values) / (n - 1)) ** 0.5