from __future__ import annotations

import io
import statistics
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

COR_PRIMARIA   = "0A7C59"
COR_LIGHT      = "E1F5EE"
COR_DARK       = "085041"
COR_ACCENT     = "1D9E75"
COR_GRAY_LIGHT = "F1EFE8"
COR_DANGER     = "A32D2D"
COR_DANGER_BG  = "FCEBEB"
COR_WARNING    = "BA7517"
COR_WARNING_BG = "FAEEDA"
COR_WHITE      = "FFFFFF"
COR_HEADER_TXT = "FFFFFF"



def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold: bool = False, size: int = 10,
        color: str = "1a1a18", italic: bool = False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _border(all_sides: bool = False) -> Border:
    thin = Side(style="thin", color="D3D1C7")
    if all_sides:
        return Border(left=thin, right=thin, top=thin, bottom=thin)
    return Border(bottom=thin)


def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> Alignment:
    return Alignment(horizontal="left", vertical="center")


def _aplicar_header_row(ws, row: int, colunas: list[str]) -> None:
    for col_idx, titulo in enumerate(colunas, start=1):
        cell = ws.cell(row=row, column=col_idx, value=titulo)
        cell.fill      = _fill(COR_PRIMARIA)
        cell.font      = _font(bold=True, color=COR_HEADER_TXT, size=10)
        cell.alignment = _center()
        cell.border    = _border(all_sides=True)


def _auto_width(ws, min_width: int = 10, max_width: int = 40) -> None:
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 3, min_width), max_width
        )


COLUNAS_HISTORICO: list[tuple[str, str]] = [
    ("ID",                   "id"),
    ("Data",                 "data"),
    ("Modalidade",           "modalidade"),
    ("Duração (min)",        "duracao_real_min"),
    ("Intensidade (1-10)",   "intensidade_percebida"),
    ("Temp (°C)",            "temperatura_c"),
    ("Umidade (%)",          "umidade_pct"),
    ("Massa Pré (kg)",       "massa_pre_kg"),
    ("Massa Pós (kg)",       "massa_pos_kg"),
    ("Ingestão (ml)",        "total_ingestao_ml"),
    ("Urina (ml)",           "total_urina_ml"),
    ("Perda Ajustada (L)",   "perda_ajustada_l"),
    ("Taxa Sudorese (L/h)",  "taxa_sudorese_lh"),
    ("Variação Massa (%)",   "variacao_massa_pct"),
    ("Balanço Hídrico (ml)", "balanco_hidrico_ml"),
    ("Recomendação (ml/h)",  "recomendacao_ml_h"),
    ("Anomalia IA",          "anomalia_detectada"),
    ("Alertas",              "total_alertas"),
]

N_COLUNAS = len(COLUNAS_HISTORICO)
ULTIMA_COLUNA = get_column_letter(N_COLUNAS) 



def _aba_historico(wb: Workbook, sessoes: list[dict], atleta: dict) -> None:
    ws = wb.active
    ws.title       = "Histórico"
    ws.freeze_panes = "A3"   

    ws.merge_cells(f"A1:{ULTIMA_COLUNA}1")
    titulo = ws["A1"]
    titulo.value = (
        f"Histórico de Hidratação — {atleta.get('codigo_anonimizado', '—')}  |  "
        f"Exportado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    titulo.fill      = _fill(COR_DARK)
    titulo.font      = _font(bold=True, color=COR_WHITE, size=12)
    titulo.alignment = _center()
    ws.row_dimensions[1].height = 28

    _aplicar_header_row(ws, 2, [c[0] for c in COLUNAS_HISTORICO])
    ws.row_dimensions[2].height = 22

    for row_idx, sessao in enumerate(sessoes, start=3):
        row_fill = _fill(COR_GRAY_LIGHT) if row_idx % 2 == 0 else _fill(COR_WHITE)

        for col_idx, (_, campo) in enumerate(COLUNAS_HISTORICO, start=1):
            valor = sessao.get(campo)

            if campo == "data" and hasattr(valor, "strftime"):
                valor = valor.strftime("%d/%m/%Y")
            elif campo == "anomalia_detectada":
                valor = "Sim" if valor else "Não"

            cell            = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.fill       = row_fill
            cell.font       = _font(size=10)
            cell.alignment  = _center()
            cell.border     = _border()

            if campo == "variacao_massa_pct" and valor is not None:
                try:
                    if float(valor) > 2:
                        cell.fill = _fill(COR_DANGER_BG)
                        cell.font = _font(bold=True, color=COR_DANGER, size=10)
                except (ValueError, TypeError):
                    pass

            if campo == "anomalia_detectada" and valor == "Sim":
                cell.fill = _fill(COR_WARNING_BG)
                cell.font = _font(bold=True, color=COR_WARNING, size=10)

    _auto_width(ws)
    ws.column_dimensions["A"].width = 6  # ID estreito



def _aba_estatisticas(wb: Workbook, sessoes: list[dict]) -> None:
    ws = wb.create_sheet("Estatísticas")

    por_modalidade: dict[str, list[float]] = {}
    for s in sessoes:
        mod  = s.get("modalidade") or "Outros"
        taxa = s.get("taxa_sudorese_lh")
        if taxa is not None:
            try:
                por_modalidade.setdefault(mod, []).append(float(taxa))
            except (ValueError, TypeError):
                pass

    ws.merge_cells("A1:G1")
    ws["A1"].value     = "Estatísticas de Taxa de Sudorese por Modalidade (L/h)"
    ws["A1"].fill      = _fill(COR_DARK)
    ws["A1"].font      = _font(bold=True, color=COR_WHITE, size=12)
    ws["A1"].alignment = _center()
    ws.row_dimensions[1].height = 26

    colunas_stats = ["Modalidade", "N Sessões", "Média", "Mediana", "Mínima", "Máxima", "Desvio Padrão"]
    _aplicar_header_row(ws, 2, colunas_stats)
    ws.row_dimensions[2].height = 20

    row = 3
    for mod, taxas in sorted(por_modalidade.items()):
        row_fill = _fill(COR_GRAY_LIGHT) if row % 2 == 0 else _fill(COR_WHITE)
        valores  = [
            mod,
            len(taxas),
            round(statistics.mean(taxas),   3),
            round(statistics.median(taxas), 3),
            round(min(taxas),               3),
            round(max(taxas),               3),
            round(statistics.stdev(taxas)   if len(taxas) > 1 else 0, 3),
        ]
        for col_idx, v in enumerate(valores, start=1):
            cell            = ws.cell(row=row, column=col_idx, value=v)
            cell.fill       = row_fill
            cell.font       = _font(size=10, bold=(col_idx == 1))
            cell.alignment  = _center() if col_idx > 1 else _left()
            cell.border     = _border()
        row += 1

    todas_taxas = [t for ts in por_modalidade.values() for t in ts]
    if todas_taxas:
        ws.row_dimensions[row].height = 20
        totais = [
            "GERAL",
            len(todas_taxas),
            round(statistics.mean(todas_taxas),   3),
            round(statistics.median(todas_taxas), 3),
            round(min(todas_taxas),               3),
            round(max(todas_taxas),               3),
            round(statistics.stdev(todas_taxas)   if len(todas_taxas) > 1 else 0, 3),
        ]
        for col_idx, v in enumerate(totais, start=1):
            cell            = ws.cell(row=row, column=col_idx, value=v)
            cell.fill       = _fill(COR_LIGHT)
            cell.font       = _font(bold=True, color=COR_DARK, size=10)
            cell.alignment  = _center() if col_idx > 1 else _left()
            cell.border     = _border(all_sides=True)

    _auto_width(ws)



def _chave_ordenacao(s: dict):
    val = s.get("criada_em") or s.get("data")
    if val is None:
        return ""
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def _aba_grafico(wb: Workbook, sessoes: list[dict]) -> None:
    ws = wb.create_sheet("Gráfico")

    ws["A1"].value = "Data"
    ws["B1"].value = "Taxa de Sudorese (L/h)"
    ws["A1"].font  = _font(bold=True)
    ws["B1"].font  = _font(bold=True)

    sessoes_ordenadas = sorted(
        [s for s in sessoes if s.get("taxa_sudorese_lh") is not None],
        key=_chave_ordenacao,
    )

    if not sessoes_ordenadas:
        ws["A2"].value = "Nenhuma sessão com taxa de sudorese disponível."
        return

    for idx, s in enumerate(sessoes_ordenadas, start=2):
        data = s.get("criada_em") or s.get("data")
        if hasattr(data, "strftime"):
            data = data.strftime("%d/%m/%Y")
        ws.cell(row=idx, column=1, value=data)
        ws.cell(row=idx, column=2, value=float(s["taxa_sudorese_lh"]))

    n        = len(sessoes_ordenadas)
    chart    = LineChart()
    chart.title         = "Evolução da Taxa de Sudorese"
    chart.y_axis.title  = "L/h"
    chart.x_axis.title  = "Sessão"
    chart.style         = 10
    chart.width         = 20
    chart.height        = 12

    data_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
    chart.add_data(data_ref, titles_from_data=True)

    serie = chart.series[0]
    serie.graphicalProperties.line.solidFill   = COR_PRIMARIA
    serie.graphicalProperties.line.width       = 20000
    serie.marker.symbol                        = "circle"
    serie.marker.size                          = 6
    serie.marker.graphicalProperties.fgColor   = COR_ACCENT
    serie.marker.graphicalProperties.solidFill = COR_ACCENT

    ws.add_chart(chart, "D2")
    _auto_width(ws)



def gerar_excel_historico(sessoes: list[dict], atleta: dict) -> bytes:
    """
    Gera planilha .xlsx com 3 abas:
    1. Histórico   — todas as sessões com destaque de alertas
    2. Estatísticas — por modalidade (média, mediana, min, max, desvio)
    3. Gráfico     — evolução da taxa de sudorese

    Args:
        sessoes: lista de dicts com dados das sessões concluídas
        atleta:  dict com ao menos 'codigo_anonimizado'

    Returns:
        bytes do arquivo .xlsx pronto para download
    """
    if not sessoes:
        raise ValueError("Nenhuma sessão fornecida para gerar o Excel.")

    wb = Workbook()
    _aba_historico(wb, sessoes, atleta)
    _aba_estatisticas(wb, sessoes)
    _aba_grafico(wb, sessoes)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()